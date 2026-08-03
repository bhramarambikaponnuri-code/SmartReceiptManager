from openpyxl import Workbook
from openpyxl.styles import Font


def export_receipt_excel(receipt_info, output_path):
    """
    Export receipt information to an Excel workbook.

    Sheet 1 : Receipt Information
    Sheet 2 : Purchased Items
    """

    wb = Workbook()

    # -------------------------------------------------
    # Receipt Information Sheet
    # -------------------------------------------------

    ws = wb.active
    ws.title = "Receipt Information"

    ws["A1"] = "Field"
    ws["B1"] = "Value"

    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)

    row = 2

    for key, value in receipt_info.items():

        if key == "Items":
            continue

        ws.cell(row=row, column=1).value = key
        ws.cell(row=row, column=2).value = value

        row += 1

    # -------------------------------------------------
    # Items Sheet
    # -------------------------------------------------

    items_sheet = wb.create_sheet("Purchased Items")

    headers = [
        "Qty",
        "Item",
        "Price",
        "Amount"
    ]

    for col, header in enumerate(headers, start=1):

        cell = items_sheet.cell(row=1, column=col)

        cell.value = header
        cell.font = Font(bold=True)

    items = receipt_info.get("Items", [])

    row = 2

    for item in items:

        items_sheet.cell(row=row, column=1).value = item.get("Qty", "")
        items_sheet.cell(row=row, column=2).value = item.get("Item", "")
        items_sheet.cell(row=row, column=3).value = item.get("Price", "")
        items_sheet.cell(row=row, column=4).value = item.get("Amount", "")

        row += 1

    wb.save(output_path)